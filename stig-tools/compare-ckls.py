#!/usr/bin/env python3
"""
compare_ckl.py

Recursively discover two CKL files, prompt user to select them by number,
then merge their SV statuses and export mismatches to an Excel workbook (xlsx).
Exports columns: Host, STIG, SV to discuss, CKL1 Status, CKL2 Status.
Adjusts column widths to fit content.
"""
import sys
import glob
import xml.etree.ElementTree as ET

# Ensure pandas/openpyxl are installed
try:
    import pandas as pd
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys as _sys
    subprocess.check_call([_sys.executable, '-m', 'pip', 'install', 'pandas', 'openpyxl'])
    import pandas as pd
    from openpyxl.utils import get_column_letter


def normalize_status(status: str) -> str:
    """
    Normalize raw STATUS into one of:
      - 'Open'
      - 'NotAFinding'
      - 'NotApplicable'
    """
    s = status.strip().lower() if status else ''
    if s.startswith('open'):
        return 'Open'
    if 'not' in s and 'finding' in s:
        return 'NotAFinding'
    return 'NotApplicable'


def parse_ckl(path: str):
    tree = ET.parse(path)
    root = tree.getroot()
    # Host
    host = root.findtext('.//HOST_NAME') or ''
    # STIG info
    stig_info = {si.findtext('SID_NAME'): si.findtext('SID_DATA')
                 for si in root.findall('.//STIG_INFO/SI_DATA')}
    title = stig_info.get('title', stig_info.get('filename', ''))
    version = stig_info.get('version', stig_info.get('ruleVersion', ''))
    revision = stig_info.get('releaseInfo', '')
    stig = title
    if version:
        stig += f" v{version}"
    if revision:
        stig += f" r{revision}"
    # Build status map
    status_map = {}
    for vuln in root.findall('.//VULN'):
        vid = None
        raw = vuln.findtext('STATUS')
        for sd in vuln.findall('STIG_DATA'):
            if sd.findtext('VULN_ATTRIBUTE') == 'Vuln_Num':
                vid = sd.findtext('ATTRIBUTE_DATA')
                break
        if not vid:
            continue
        status_map[vid] = (normalize_status(raw), raw, vuln)
    return status_map, host, stig


def choose_ckls():
    files = sorted(glob.glob('**/*.ckl', recursive=True))
    if len(files) < 2:
        print('Error: require at least two CKL files.', file=sys.stderr)
        sys.exit(1)
    print('Available CKL files:')
    for i, f in enumerate(files, 1):
        print(f'  {i}) {f}')
    def get_choice(prompt):
        while True:
            try:
                n = int(input(prompt))
                if 1 <= n <= len(files):
                    return files[n-1]
            except ValueError:
                pass
            print('Invalid choice, try again.')
    return get_choice('Select FIRST CKL by number: '), get_choice('Select SECOND CKL by number: ')


def main():
    ckl1_path, ckl2_path = choose_ckls()
    m1, host, stig = parse_ckl(ckl1_path)
    m2, _, _ = parse_ckl(ckl2_path)

    rows = []
    for vid in sorted(set(m1) | set(m2)):
        s1, raw1, _ = m1.get(vid, ('', None, None))
        s2, raw2, _ = m2.get(vid, ('', None, None))
        # skip if both missing or same raw
        if not raw1 and not raw2:
            continue
        if raw1 and raw2 and raw1 == raw2:
            continue
        # Only Open vs NotAFinding pairs
        if not ((s1 == 'Open' and s2 == 'NotAFinding') or (s1 == 'NotAFinding' and s2 == 'Open')):
            continue
        rows.append({
            'Host': host,
            'STIG': stig,
            'SV to discuss': vid,
            'CKL1 Status': s1,
            'CKL2 Status': s2
        })

    if not rows:
        print('No deltas found.')
        sys.exit(0)

    df = pd.DataFrame(rows)
    # Sort: Open first in CKL1
    prio = {'Open': 1, 'NotAFinding': 2}
    df['prio'] = df['CKL1 Status'].map(lambda x: prio.get(x, 99))
    df.sort_values(['prio', 'SV to discuss'], inplace=True)
    df.drop(columns=['prio'], inplace=True)

    out = 'deltas.xlsx'
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Deltas')
        ws = writer.sheets['Deltas']
        for i, col in enumerate(df.columns, 1):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[get_column_letter(i)].width = max_len
    print(f'Exported {len(df)} deltas to {out}')


if __name__ == '__main__':
    main()
